import win32com.client
from openpyxl import load_workbook


class Macro_ReplaceValues_Selection:
    def __init__(self):
        self.word_app = win32com.client.Dispatch('Word.Application')
        self.excel_file = None

    def load_excel_file(self, excel_file_path):
        self.excel_file = load_workbook(excel_file_path)

    def replace_values(self, use_wildcards):
        try:
            if self.word_app is None:
                return "Word is not open."
            if self.excel_file is None:
                return "No Excel file loaded."

            # Get the current selection
            selection = self.word_app.Selection
            # If no text is selected, there's nothing to do
            if not selection.Text:
                raise Exception("No text is selected in Word.")

            excel_sheet = self.excel_file.active
            replacements = {}
            for row in excel_sheet.iter_rows():
                find_text = str(row[0].value)
                # If the cell is empty, stop the loop
                if find_text is None:
                    break
                replace_text = str(row[1].value)
                if find_text and replace_text:
                    if use_wildcards:
                        # Convert wildcards to regex
                        find_text = find_text.replace("~*", "\\*")
                        find_text = find_text.replace("~?", "\\?")
                        find_text = find_text.replace("~~", "~")
                        find_text = find_text.replace("?", ".")
                        find_text = find_text.replace("*", ".*")
                        find_text = find_text.replace("[!", "[^")
                        find_text = find_text.replace("#", "\\d")
                    replacements[find_text] = replace_text

            # Iterate through the selection and replace words
            for find_text, replace_text in replacements.items():
                # Word's Find and Replace function
                find = selection.Find
                find.ClearFormatting()
                find.Replacement.ClearFormatting()
                find.Text = find_text
                find.Replacement.Text = replace_text
                find.Forward = True
                find.Wrap = win32com.client.constants.wdFindStop
                find.Format = False
                find.MatchCase = True  # Make it case-sensitive
                find.MatchWholeWord = False
                find.MatchWildcards = use_wildcards
                find.MatchSoundsLike = False
                find.MatchAllWordForms = False
                find.Execute(Replace=win32com.client.constants.wdReplaceAll)

        except Exception as e:
            return str(e)

    def save_document(self):
        if self.word_app is None:
            raise Exception("Word is not open.")

        self.word_app.ActiveDocument.Save()
